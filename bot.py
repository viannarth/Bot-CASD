import openpyxl
from urllib.parse import quote
import webbrowser
from time import sleep
import pyautogui
from unidecode import unidecode
import re

BIMESTRE = # Insira o bimestre atual aqui
SEMANA = # Insira a semana atual aqui

# Abertura prévia do WhatsApp Web
webbrowser.open('https://web.whatsapp.com/')
sleep(60)  # Aguarda o usuário realizar a autenticação
# Após a autenticação inicial, essas duas linhas podem ser comentadas, para agilizar o processo

# Função que formata o numero do telefone(insere o código do país)
def formatar_numero(numero):
    return f"55{numero}"

# Map que associa um nome a um numero. Por exemplo: dados_formatados[jonathan]=5521984334015
dados_formatados = {}

# Abrir a planilha e realizar a leitura dos dados
workbook_1 = openpyxl.load_workbook('Controle de Presença 2024.xlsx')
página_telefones = workbook_1['N° de telefone']

for linha in página_telefones.iter_rows(min_row=2):   
    nome = linha[0].value
    numero = str(linha[1].value)  # Garantir que o número seja uma string
    numero_formatado = formatar_numero(numero)
    dados_formatados[nome] = numero_formatado

# Vetor para guardar o nome dos desistentes e, assim, excluí-los do processo de automação.
nome_desistentes = []

# Leitura da planilha que contém as desistências e armazenamento no vetor nome_desistentes
workbook_2 = openpyxl.load_workbook('Controle de Presença 2024.xlsx')
página_desistentes = workbook_2['Desistências']

for linha in página_desistentes.iter_rows(min_row=1):
    nome = unidecode(str(linha[0].value).upper()) # Colocar todas as letras maiúsculas e retirar os acentos
    nome_sem_espacos = re.sub(r'\s+', ' ', nome).strip() # Função que substitui todos os espaços duplicados por um único
    nome_desistentes.append(nome_sem_espacos)

# Lógica de contadores para atualizar o nome da planilha conforme o bimestre
nome_aba = f'{BIMESTRE}° Bimestre (extensivo+semi)'

# Leitura da planilha que contém os nomes dos alunos, semanas, dias, suas faltas em cada dia e em cada semana
workbook_3 = openpyxl.load_workbook('Controle de Presença 2024.xlsx')
página_alunos = workbook_3[nome_aba]


for linha in página_alunos.iter_rows(min_row=4):
    nome = linha[0].value

    # Verificar se o nome do aluno na lista geral também está na lista de desistentes.
    if nome in nome_desistentes:
        continue  # Se sim, pula para a próxima iteração.

    # Coletar o numero de celular do aluno em dados_formatados
    numero_1 = dados_formatados.get(nome)
    if not numero_1:
        continue  # Caso não haja número de celular para o aluno fornecido na planilha

    cont = 0
    dias_faltas = []
    turma = linha[1].value
    seg = linha[2 + 7 *(SEMANA-1)].value
    ter = linha[3 + 7 * (SEMANA-1)].value
    quar = linha[4 + 7 * (SEMANA-1)].value
    quin = linha[5 + 7 * (SEMANA-1)].value
    sex = linha[6 + 7 * (SEMANA-1)].value

    if seg == 'F':
        cont += 1
        dias_faltas.append('Segunda-feira')
    if ter == 'F':
        cont += 1
        dias_faltas.append('Terça-feira')
    if quar == 'F':
        cont += 1
        dias_faltas.append('Quarta-feira')
    if quin == 'F':
        cont += 1
        dias_faltas.append('Quinta-feira')
    if sex == 'F':
        cont += 1
        dias_faltas.append('Sexta-feira')

    faltas = cont

    # Formatar a lista de dias faltados em uma string
    dias_faltados_str = ', '.join(dias_faltas)

    # Selecionar a mensagem com base no número de faltas
    if faltas == 1:
        mensagem = f'Olá, {nome}. Você teve uma falta nessa semana ({SEMANA}). Faltou no seguinte dia: {dias_faltados_str}. Por favor, justifique através do seguinte formulário: https://forms.gle/dg7LPGudht6XKGVv8'
    elif faltas == 2:
        mensagem = f'Olá, {nome}. Estamos contatando você para dizer que tem duas faltas nessa semana ({SEMANA}). Faltou nos seguintes dias: {dias_faltados_str}. Por favor, justifique através do seguinte formulário: https://forms.gle/dg7LPGudht6XKGVv8'
    elif faltas == 3:
        mensagem = f'Olá, {nome}. Tudo bem com você? Notamos que você teve 3 faltas nessa semana ({SEMANA}) nos dias: {dias_faltados_str}. Explique sua situação para que compreendamos seus motivos através do seguinte formulário: https://forms.gle/dg7LPGudht6XKGVv8'
    elif faltas == 4:
        mensagem = f'Olá, {nome}. Tudo bem com você? Então, percebemos que você teve 4 faltas nessa semana ({SEMANA}) nos dias: {dias_faltados_str}. Explique para nós sua situação para que possamos compreender suas causas e talvez ajudá-lo(a) através do seguinte formulário: https://forms.gle/dg7LPGudht6XKGVv8'
    elif faltas == 5:
        mensagem = f'Olá, {nome}. Tudo bem com você? Estamos enviando essa mensagem pois percebemos que você faltou a semana ({SEMANA}) inteira. Caso possamos fazer alguma coisa para ajudar, gostaríamos de saber como está o seu dia a dia e se está tendo algum problema para frequentar as aulas. Deixe sua resposta no seguinte formulário: https://forms.gle/dg7LPGudht6XKGVv8'
    else:
        continue
    #Função erro para filtrar os numeros e nomes dos alunos para os quais a mensagem não conseguiu ser enviada para,posteriormente, tentar o reenvio.
    try:
        # Codificar a mensagem para o formato URL
        mensagem_codificada = quote(mensagem)
        link_mensagem_whatsapp = f'https://web.whatsapp.com/send?phone={numero_1}&text={mensagem_codificada}'

        # Abrir o link no navegador
        webbrowser.open(link_mensagem_whatsapp)
        sleep(10)  # Tempo de espera para garantir que o WhatsappWeb processe as mensagens.

        # Envio da mensagem
        pyautogui.hotkey('enter')
        sleep(3)
        # Fechar guia do WhatsApp
        pyautogui.hotkey('ctrl', 'w')
        sleep(3)

    except Exception as e:
        print(f'Não foi possivel enviar a mensagem para {nome}')
        with open('erros_de_envio.txt', 'a', newline='', encoding='utf-8') as arquivo:
            arquivo.write(f'{nome}, {numero_1}\n')
